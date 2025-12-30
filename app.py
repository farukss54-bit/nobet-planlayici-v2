"""
Nöbet Planlayıcı - Ana Uygulama

Streamlit tabanlı kullanıcı arayüzü.
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Yerel modüller
from models import Ayarlar, Personel, EslesmeTercihi, AylikPlan, Alan, KidemGrubu, VardiyaTipi, HAZIR_VARDIYALAR
from utils import (
    ay_gun_sayisi, resmi_tatiller, gun_parse, 
    hafta_gunu_adi, tum_hafta_gunleri, hafta_gunu_numarasi
)
from storage import (
    ayarlari_kaydet, ayarlari_yukle_veya_varsayilan,
    aylik_plani_kaydet, aylik_plani_yukle_veya_yeni,
    kayitli_planlari_listele, ayarlari_json_olarak_export,
    ayarlari_json_dan_import
)
from solver import NobetSolver, SolverInput, SolverConfig, AlanTanimi, VardiyaTanimi, cozum_bulunamadi_teshis

# Demo senaryo modülü
from streamlit_integration import (
    get_demo_sidebar,
    render_demo_detail_modal,
    is_demo_active,
    get_demo_meta
)


# =============================================================================
# SAYFA AYARLARI
# =============================================================================

st.set_page_config(page_title="Nöbet Planlayıcı", layout="wide")
st.title("🏥 Acil Servis Nöbet Planlayıcı")

# Demo modu aktifse detaylı özet göster
if is_demo_active():
    meta = get_demo_meta()
    
    with st.expander(f"🧪 **Demo Modu Aktif** - {meta.get('difficulty', '?')} | Seed: {meta.get('seed', '?')}", expanded=True):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("👥 Personel", len(st.session_state.get('personel_list', [])))
        
        with col2:
            izin_toplam = sum(len(v) for v in st.session_state.get('izin_map', {}).values())
            st.metric("🏖️ Toplam İzin", izin_toplam)
        
        with col3:
            kisit_toplam = (
                len(st.session_state.get('no_pairs_list', [])) +
                len(st.session_state.get('soft_no_pairs_list', []))
            )
            st.metric("🚫 Çift Kısıtları", kisit_toplam)
        
        with col4:
            alan_sayisi = len(st.session_state.get('alanlar', []))
            vardiya_sayisi = len(st.session_state.get('vardiya_tipleri', []))
            st.metric("🏢/⏰ Alan/Vardiya", f"{alan_sayisi}/{vardiya_sayisi}")
        
        # Kapasite/Hedef hesapla ve göster
        alanlar = st.session_state.get('alanlar', [])
        vardiyalar = st.session_state.get('vardiya_tipleri', [])
        gun_sayisi = meta.get('gun_sayisi', 30)
        
        if alanlar:
            toplam_kontenjan = sum(a.get('kontenjan', 1) for a in alanlar)
        else:
            toplam_kontenjan = 1
        
        if vardiyalar:
            gunluk_slot = toplam_kontenjan * len(vardiyalar)
        else:
            gunluk_slot = toplam_kontenjan
        
        demo_kapasite = gunluk_slot * gun_sayisi
        
        st.caption(f"📅 Dönem: {meta.get('yil', '?')}-{meta.get('ay', '?'):02d} | 📊 Demo Kapasite: {demo_kapasite} | ✅ Çözüm sekmesine git")

# Demo detay modalı
render_demo_detail_modal()


# =============================================================================
# SESSION STATE BAŞLATMA
# =============================================================================

def init_session_state():
    """Session state'i başlat veya kayıtlı ayarları yükle"""
    
    # Demo modu aktifse ASLA kayıtlı dosyadan yükleme - demo verisi kullanılacak
    if st.session_state.get("_demo_aktif", False):
        # Demo verisi zaten session_state'te, sadece initialized flag'i set et
        st.session_state["initialized"] = True
        return
    
    if "initialized" not in st.session_state:
        # Kayıtlı ayarları yükle
        ayarlar = ayarlari_yukle_veya_varsayilan()
        
        # Personel listesi
        if ayarlar.personeller:
            st.session_state["personel_list"] = [p.isim for p in ayarlar.personeller]
            st.session_state["personel_targets"] = {
                p.isim: p.hedef_nobet 
                for p in ayarlar.personeller 
                if p.hedef_nobet is not None
            }
            st.session_state["weekday_block_map"] = {
                p.isim: p.bloklu_gunler 
                for p in ayarlar.personeller
            }
        else:
            st.session_state["personel_list"] = [
                "Dr. Ahmet", "Dr. Ayşe", "Dr. Mehmet", "Dr. Fatma", 
                "Dr. Ali", "Dr. Zeynep", "Dr. Can", "Dr. Elif", "Dr. Burak"
            ]
            st.session_state["personel_targets"] = {}
            st.session_state["weekday_block_map"] = {}
        
        st.session_state["personel_sayisi"] = len(st.session_state["personel_list"])
        
        # Eşleşme kuralları
        st.session_state["want_pairs_list"] = [
            {"a": e.personel_a, "b": e.personel_b, "min": e.min_birlikte}
            for e in ayarlar.birlikte_tutma
        ]
        st.session_state["no_pairs_list"] = [
            {"a": e.personel_a, "b": e.personel_b}
            for e in ayarlar.ayri_tutma
        ]
        st.session_state["soft_no_pairs_list"] = [
            {"a": e.personel_a, "b": e.personel_b}
            for e in ayarlar.esnek_ayri_tutma
        ]
        
        # Ağırlıklar
        st.session_state["varsayilan_hedef"] = ayarlar.varsayilan_hedef
        
        # Tarih (varsayılan: gelecek ay)
        bugun = datetime.now()
        if bugun.month == 12:
            st.session_state["yil"] = bugun.year + 1
            st.session_state["ay"] = 1
        else:
            st.session_state["yil"] = bugun.year
            st.session_state["ay"] = bugun.month + 1
        
        # Ay'a özel veriler
        st.session_state["izin_map"] = {}
        st.session_state["prefer_map"] = {}
        st.session_state["manuel_tatiller"] = ""
        
        # Aşama 1: Alanlar
        st.session_state["alanlar"] = [
            {"isim": a.isim, "kontenjan": a.gunluk_kontenjan, "max_kontenjan": a.max_kontenjan, "renk": a.renk}
            for a in ayarlar.alanlar
        ] if ayarlar.alanlar else []
        st.session_state["alan_modu_aktif"] = len(st.session_state["alanlar"]) > 0
        st.session_state["alan_bazli_denklik"] = ayarlar.alan_bazli_denklik
        
        # Personel alan yetkinlikleri
        st.session_state["personel_alan_yetkinlikleri"] = {
            p.isim: p.calisabilir_alanlar
            for p in ayarlar.personeller
            if p.calisabilir_alanlar
        }
        
        # Kıdem grupları
        st.session_state["kidem_gruplari"] = [
            {"isim": k.isim, "renk": k.renk, "varsayilan_hedef": k.varsayilan_hedef}
            for k in ayarlar.kidem_gruplari
        ] if ayarlar.kidem_gruplari else []
        
        st.session_state["personel_kidem_gruplari"] = {
            p.isim: p.kidem_grubu
            for p in ayarlar.personeller
            if p.kidem_grubu
        }
        
        # Vardiya tipleri
        st.session_state["vardiya_tipleri"] = [
            {"isim": v.isim, "baslangic": v.baslangic, "bitis": v.bitis, "renk": v.renk}
            for v in ayarlar.vardiya_tipleri
        ] if ayarlar.vardiya_tipleri else []
        
        st.session_state["personel_vardiya_kisitlari"] = {
            p.isim: p.calisabilir_vardiyalar
            for p in ayarlar.personeller
            if p.calisabilir_vardiyalar
        }
        
        st.session_state["saat_bazli_denge"] = ayarlar.saat_bazli_denge
        
        # Kural ayarları
        st.session_state["ardisik_yasak"] = ayarlar.ardisik_yasak
        st.session_state["gunasiri_limit_aktif"] = ayarlar.gunasiri_limit_aktif
        st.session_state["max_gunasiri"] = ayarlar.max_gunasiri
        st.session_state["hafta_sonu_dengesi"] = ayarlar.hafta_sonu_dengesi
        st.session_state["w_cuma"] = ayarlar.w_cuma
        st.session_state["w_cumartesi"] = ayarlar.w_cumartesi
        st.session_state["w_pazar"] = ayarlar.w_pazar
        st.session_state["tatil_dengesi"] = ayarlar.tatil_dengesi
        st.session_state["iki_gun_bosluk_aktif"] = ayarlar.iki_gun_bosluk_aktif
        st.session_state["w_gap3"] = ayarlar.iki_gun_bosluk_tercihi
        
        st.session_state["initialized"] = True


def session_to_ayarlar() -> Ayarlar:
    """Session state'ten Ayarlar nesnesi oluşturur"""
    personeller = []
    for isim in st.session_state.get("personel_list", []):
        personeller.append(Personel(
            isim=isim,
            hedef_nobet=st.session_state.get("personel_targets", {}).get(isim),
            bloklu_gunler=st.session_state.get("weekday_block_map", {}).get(isim, []),
            calisabilir_alanlar=st.session_state.get("personel_alan_yetkinlikleri", {}).get(isim, []),
            kidem_grubu=st.session_state.get("personel_kidem_gruplari", {}).get(isim),
            calisabilir_vardiyalar=st.session_state.get("personel_vardiya_kisitlari", {}).get(isim, [])
        ))
    
    birlikte_tutma = [
        EslesmeTercihi(
            personel_a=item["a"],
            personel_b=item["b"],
            min_birlikte=item.get("min", 0)
        )
        for item in st.session_state.get("want_pairs_list", [])
    ]
    
    ayri_tutma = [
        EslesmeTercihi(personel_a=item["a"], personel_b=item["b"])
        for item in st.session_state.get("no_pairs_list", [])
    ]
    
    esnek_ayri_tutma = [
        EslesmeTercihi(personel_a=item["a"], personel_b=item["b"], zorunlu=False)
        for item in st.session_state.get("soft_no_pairs_list", [])
    ]
    
    # Alanlar
    alanlar = [
        Alan(
            isim=a["isim"],
            gunluk_kontenjan=a.get("kontenjan", 1),
            max_kontenjan=a.get("max_kontenjan"),
            renk=a.get("renk", "#808080"),
            kidem_kurallari=a.get("kidem_kurallari", {})
        )
        for a in st.session_state.get("alanlar", [])
    ]
    
    # Kıdem grupları
    kidem_gruplari = [
        KidemGrubu(
            isim=k["isim"], 
            renk=k.get("renk", "#808080"),
            varsayilan_hedef=k.get("varsayilan_hedef")
        )
        for k in st.session_state.get("kidem_gruplari", [])
    ]
    
    # Vardiya tipleri
    vardiya_tipleri = [
        VardiyaTipi(
            isim=v["isim"],
            baslangic=v.get("baslangic", "08:00"),
            bitis=v.get("bitis", "16:00"),
            renk=v.get("renk", "#808080")
        )
        for v in st.session_state.get("vardiya_tipleri", [])
    ]
    
    return Ayarlar(
        personeller=personeller,
        varsayilan_hedef=st.session_state.get("varsayilan_hedef", 7),
        alanlar=alanlar,
        alan_bazli_denklik=st.session_state.get("alan_bazli_denklik", True),
        kidem_gruplari=kidem_gruplari,
        vardiya_tipleri=vardiya_tipleri,
        saat_bazli_denge=st.session_state.get("saat_bazli_denge", True),
        birlikte_tutma=birlikte_tutma,
        ayri_tutma=ayri_tutma,
        esnek_ayri_tutma=esnek_ayri_tutma,
        # Kural ayarları
        ardisik_yasak=st.session_state.get("ardisik_yasak", True),
        gunasiri_limit_aktif=st.session_state.get("gunasiri_limit_aktif", True),
        max_gunasiri=st.session_state.get("max_gunasiri", 1),
        hafta_sonu_dengesi=st.session_state.get("hafta_sonu_dengesi", True),
        w_cuma=st.session_state.get("w_cuma", 1000),
        w_cumartesi=st.session_state.get("w_cumartesi", 1000),
        w_pazar=st.session_state.get("w_pazar", 1000),
        tatil_dengesi=st.session_state.get("tatil_dengesi", True),
        iki_gun_bosluk_aktif=st.session_state.get("iki_gun_bosluk_aktif", True),
        iki_gun_bosluk_tercihi=st.session_state.get("w_gap3", 300)
    )


init_session_state()


# =============================================================================
# SIDEBAR - KAYDETME/YÜKLEME
# =============================================================================

with st.sidebar:
    st.header("💾 Veri Yönetimi")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("💾 Kaydet", use_container_width=True, help="Ayarları kaydet"):
            ayarlar = session_to_ayarlar()
            if ayarlari_kaydet(ayarlar):
                st.success("✓ Kaydedildi")
            else:
                st.error("Kaydetme hatası")
    
    with col2:
        if st.button("🔄 Yükle", use_container_width=True, help="Kayıtlı ayarları yükle"):
            st.session_state.clear()
            st.rerun()
    
    st.divider()
    
    # JSON Export/Import
    with st.expander("📤 Dışa/İçe Aktar"):
        ayarlar = session_to_ayarlar()
        json_str = ayarlari_json_olarak_export(ayarlar)
        
        st.download_button(
            "⬇️ Ayarları İndir (JSON)",
            data=json_str,
            file_name="nobet_ayarlari.json",
            mime="application/json"
        )
        
        uploaded = st.file_uploader("Ayar dosyası yükle", type=["json"])
        if uploaded:
            content = uploaded.read().decode("utf-8")
            loaded = ayarlari_json_dan_import(content)
            if loaded:
                ayarlari_kaydet(loaded)
                st.success("Ayarlar yüklendi! Sayfayı yenileyin.")
                if st.button("🔄 Yenile"):
                    st.session_state.clear()
                    st.rerun()
    
    st.divider()
    
    # Geçmiş planlar
    st.subheader("📅 Geçmiş Planlar")
    planlar = kayitli_planlari_listele()
    
    if planlar:
        for plan in planlar[:5]:  # Son 5 plan
            ay_adi = datetime(plan["yil"], plan["ay"], 1).strftime("%B %Y")
            durum = "✓" if plan["sonuc_var"] else "○"
            st.caption(f"{durum} {ay_adi}")
    else:
        st.caption("Henüz kaydedilmiş plan yok")
    
    # Demo Senaryo Kontrolleri
    get_demo_sidebar()


# =============================================================================
# ANA SEKMELER
# =============================================================================

tabs = st.tabs(["👥 Kişiler", "🎖️ Kıdem", "🏢 Alanlar", "⏰ Vardiyalar", "🏖️ İzinler", "👫 Eşleşmeler", "✅ Çözüm"])


# =============================================================================
# TAB 0: KİŞİLER
# =============================================================================

with tabs[0]:
    st.subheader("👥 Kişiler ve Hedefler")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.number_input("Yıl", min_value=2020, max_value=2100, step=1, key="yil")
    with col2:
        st.number_input("Ay", min_value=1, max_value=12, step=1, key="ay")
    with col3:
        # Değeri 0-31 arasına sınırla
        current_hedef = st.session_state.get("varsayilan_hedef", 7)
        clamped_hedef = max(0, min(31, current_hedef))
        if current_hedef != clamped_hedef:
            st.session_state["varsayilan_hedef"] = clamped_hedef
        st.number_input(
            "Varsayılan hedef nöbet", 
            min_value=0, max_value=31, step=1, 
            key="varsayilan_hedef"
        )
    
    st.divider()
    
    # Personel sayısı
    personel_sayisi = st.number_input(
        "Kaç personel var?",
        min_value=1, max_value=50,
        value=st.session_state.get("personel_sayisi", 9),
        step=1,
        key="personel_sayisi_input"
    )
    
    # Listeyi güncelle
    current_list = st.session_state["personel_list"]
    if len(current_list) < personel_sayisi:
        for i in range(len(current_list), personel_sayisi):
            current_list.append(f"Personel {i+1}")
    elif len(current_list) > personel_sayisi:
        st.session_state["personel_list"] = current_list[:personel_sayisi]
    
    st.session_state["personel_sayisi"] = personel_sayisi
    
    st.caption("Her personelin adını ve hedef nöbet sayısını girin:")
    
    default_target = st.session_state.get("varsayilan_hedef", 7)
    
    for i in range(personel_sayisi):
        cols = st.columns([3, 1])
        with cols[0]:
            st.session_state["personel_list"][i] = st.text_input(
                f"{i+1}. Personel",
                value=st.session_state["personel_list"][i],
                key=f"personel_name_{i}"
            )
        with cols[1]:
            p_name = st.session_state["personel_list"][i]
            current_target = st.session_state.get("personel_targets", {}).get(p_name, default_target)
            new_target = st.number_input(
                "Hedef",
                min_value=0, max_value=31,
                value=int(current_target),
                step=1,
                key=f"target_{i}"
            )
            if new_target != default_target:
                st.session_state.setdefault("personel_targets", {})[p_name] = new_target
            elif p_name in st.session_state.get("personel_targets", {}):
                st.session_state["personel_targets"].pop(p_name, None)


# =============================================================================
# TAB 1: KIDEM GRUPLARI (YENİ)
# =============================================================================

with tabs[1]:
    st.subheader("🎖️ Kıdem Grupları")
    
    st.info("""
    **Kıdem Grupları**: Personeli gruplara ayırabilirsiniz (örn: Asistan, Uzman, Profesör).
    Her grup için varsayılan nöbet sayısı belirleyebilir, alan bazlı kurallar tanımlayabilirsiniz.
    """)
    
    # ====== GRUP TANIMLAMA ======
    st.markdown("### ➕ Grup Ekle")
    col1, col2, col3 = st.columns([3, 1, 1])
    
    with col1:
        yeni_grup_isim = st.text_input("Grup adı", placeholder="Örn: Uzman", key="yeni_grup_isim")
    with col2:
        yeni_grup_hedef = st.number_input("Hedef nöbet", min_value=0, max_value=31, value=7, key="yeni_grup_hedef", help="Bu gruptaki personelin varsayılan aylık nöbet sayısı")
    with col3:
        yeni_grup_renk = st.color_picker("Renk", value="#4CAF50", key="yeni_grup_renk")
    
    if st.button("➕ Grup Ekle", key="grup_ekle_btn"):
        if yeni_grup_isim.strip():
            mevcut_gruplar = st.session_state.get("kidem_gruplari", [])
            mevcut_isimler = [g["isim"] for g in mevcut_gruplar]
            if yeni_grup_isim.strip() not in mevcut_isimler:
                st.session_state.setdefault("kidem_gruplari", []).append({
                    "isim": yeni_grup_isim.strip(),
                    "renk": yeni_grup_renk,
                    "varsayilan_hedef": yeni_grup_hedef
                })
                st.rerun()
            else:
                st.error("Bu isimde bir grup zaten var!")
        else:
            st.error("Grup adı boş olamaz!")
    
    st.divider()
    
    # ====== MEVCUT GRUPLAR ======
    st.markdown("### 📋 Mevcut Gruplar")
    
    kidem_gruplari = st.session_state.get("kidem_gruplari", [])
    vardiyalar = st.session_state.get("vardiya_tipleri", [])
    
    if not kidem_gruplari:
        st.caption("Henüz kıdem grubu tanımlanmamış.")
    else:
        for i, grup in enumerate(kidem_gruplari):
            # Gruptaki kişi sayısını hesapla
            personel_gruplari = st.session_state.get("personel_kidem_gruplari", {})
            kisi_sayisi = sum(1 for g in personel_gruplari.values() if g == grup["isim"])
            
            with st.expander(
                f"● {grup['isim']} ({kisi_sayisi} kişi)", 
                expanded=False
            ):
                col1, col2, col3 = st.columns([2, 2, 1])
                
                with col1:
                    st.markdown(
                        f"<span style='color:{grup.get('renk', '#808080')}; font-size:24px'>●</span>", 
                        unsafe_allow_html=True
                    )
                    yeni_renk = st.color_picker(
                        "Renk",
                        value=grup.get("renk", "#808080"),
                        key=f"grup_renk_{i}"
                    )
                    if yeni_renk != grup.get("renk"):
                        st.session_state["kidem_gruplari"][i]["renk"] = yeni_renk
                
                with col3:
                    if st.button("🗑️ Grubu Sil", key=f"grup_sil_{i}"):
                        # Önce bu gruptaki personellerin atamalarını kaldır
                        for p, g in list(personel_gruplari.items()):
                            if g == grup["isim"]:
                                del personel_gruplari[p]
                        st.session_state["kidem_gruplari"].pop(i)
                        st.rerun()
                
                st.divider()
                
                # VARDIYA BAZLI HEDEF veya TOPLAM HEDEF
                if vardiyalar:
                    st.markdown("**Vardiya Bazlı Hedefler:**")
                    
                    vardiya_hedefleri = grup.get("vardiya_hedefleri", {})
                    toplam_nobet = 0
                    toplam_saat = 0
                    
                    # Her vardiya için hedef input
                    vcols = st.columns(min(len(vardiyalar), 4))
                    for v_idx, v in enumerate(vardiyalar):
                        with vcols[v_idx % 4]:
                            mevcut = vardiya_hedefleri.get(v["isim"], 0)
                            yeni = st.number_input(
                                f"{v['isim']}",
                                min_value=0, max_value=31,
                                value=int(mevcut),
                                key=f"grup_{i}_vardiya_{v_idx}",
                                help=f"Bu gruptan {v['isim']} vardiyasında kaç nöbet"
                            )
                            
                            # Güncelle
                            if yeni != mevcut:
                                if "vardiya_hedefleri" not in st.session_state["kidem_gruplari"][i]:
                                    st.session_state["kidem_gruplari"][i]["vardiya_hedefleri"] = {}
                                st.session_state["kidem_gruplari"][i]["vardiya_hedefleri"][v["isim"]] = yeni
                            
                            # Saat hesapla
                            try:
                                from models import VardiyaTipi
                                vt = VardiyaTipi(v["isim"], v.get("baslangic", "08:00"), v.get("bitis", "08:00"))
                                toplam_saat += yeni * vt.saat
                            except:
                                toplam_saat += yeni * 24
                            toplam_nobet += yeni
                    
                    st.caption(f"📊 Toplam: {toplam_nobet} nöbet, {toplam_saat} saat")
                    
                    # Eski hedefi de güncelle (uyumluluk için)
                    st.session_state["kidem_gruplari"][i]["varsayilan_hedef"] = toplam_nobet
                    
                else:
                    # Vardiya tanımlı değilse eski mod
                    st.markdown("**Toplam Nöbet Hedefi:**")
                    mevcut_hedef = grup.get("varsayilan_hedef", 7)
                    yeni_hedef = st.number_input(
                        "Hedef nöbet sayısı",
                        min_value=0, max_value=31,
                        value=mevcut_hedef if mevcut_hedef else 7,
                        key=f"grup_hedef_{i}"
                    )
                    if yeni_hedef != mevcut_hedef:
                        st.session_state["kidem_gruplari"][i]["varsayilan_hedef"] = yeni_hedef
                    
                    st.caption("💡 Vardiyalar sekmesinden vardiya tanımlarsanız, vardiya bazlı hedef girebilirsiniz.")
    
    st.divider()
    
    # ====== GRUPLARA PERSONEL ATAMA ======
    st.markdown("### 👤 Gruplara Personel Atama")
    
    personeller = st.session_state.get("personel_list", [])
    
    if not personeller:
        st.warning("Önce Kişiler sekmesinde personel ekleyin.")
    elif not kidem_gruplari:
        st.warning("Önce yukarıda kıdem grupları tanımlayın.")
    else:
        st.caption("Her grup için personel seçin. Bir personel sadece bir grupta olabilir.")
        
        personel_kidem = st.session_state.get("personel_kidem_gruplari", {})
        
        # Personel listesinde olmayan atamaları temizle (demo değişikliği için)
        gecersiz_atamalar = [p for p in personel_kidem.keys() if p not in personeller]
        for p in gecersiz_atamalar:
            del personel_kidem[p]
        
        for i, grup in enumerate(kidem_gruplari):
            grup_isim = grup["isim"]
            grup_renk = grup.get("renk", "#808080")
            
            # Bu grupta olan personeller (sadece mevcut personel listesinde olanlar)
            mevcut_uyeler = [p for p, g in personel_kidem.items() if g == grup_isim and p in personeller]
            
            # Başka gruplarda olmayan personeller (seçilebilir)
            baska_gruplarda = [p for p, g in personel_kidem.items() if g != grup_isim]
            musait_personeller = [p for p in personeller if p not in baska_gruplarda or p in mevcut_uyeler]
            
            # Default değerlerin options'da olduğundan emin ol
            valid_defaults = [p for p in mevcut_uyeler if p in musait_personeller]
            
            with st.expander(f"● {grup_isim} ({len(valid_defaults)} kişi)", expanded=len(valid_defaults) == 0):
                secilen = st.multiselect(
                    f"Personel seç",
                    options=musait_personeller,
                    default=valid_defaults,
                    key=f"grup_personel_{i}",
                    label_visibility="collapsed"
                )
                
                # Atamaları güncelle
                # Önce bu gruptan çıkarılanları temizle
                for p in mevcut_uyeler:
                    if p not in secilen:
                        if p in st.session_state.get("personel_kidem_gruplari", {}):
                            del st.session_state["personel_kidem_gruplari"][p]
                
                # Yeni eklenenler
                for p in secilen:
                    st.session_state.setdefault("personel_kidem_gruplari", {})[p] = grup_isim
        
        # Atanmamış personelleri göster
        atanmamis = [p for p in personeller if p not in personel_kidem]
        if atanmamis:
            st.info(f"⚠️ Atanmamış personeller ({len(atanmamis)}): {', '.join(atanmamis)}")


# =============================================================================
# TAB 2: ALANLAR
# =============================================================================

with tabs[2]:
    st.subheader("🏢 Çalışma Alanları")
    
    st.info("""
    **Çoklu Alan Modu**: Farklı çalışma alanları tanımlayabilirsiniz (örn: Yeşil, Sarı, Kırmızı alan).
    Her alana günlük kontenjan belirlenir ve sistem otomatik olarak dağılımı yapar.
    
    Alan tanımlamazsanız, mevcut tek-alan modu kullanılır.
    """)
    
    # Alan modu toggle
    alan_modu = st.checkbox(
        "Çoklu alan modunu aktifleştir",
        value=st.session_state.get("alan_modu_aktif", False),
        key="alan_modu_checkbox"
    )
    st.session_state["alan_modu_aktif"] = alan_modu
    
    if alan_modu:
        st.divider()
        
        # Yeni alan ekleme
        st.markdown("### ➕ Alan Ekle")
        col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
        
        with col1:
            yeni_alan_isim = st.text_input("Alan adı", placeholder="Örn: Kırmızı Alan", key="yeni_alan_isim")
        with col2:
            yeni_alan_kontenjan = st.number_input("Hedef", min_value=1, max_value=10, value=1, key="yeni_alan_kont", help="Günlük hedef kişi sayısı")
        with col3:
            yeni_alan_max = st.number_input("Max", min_value=1, max_value=15, value=3, key="yeni_alan_max", help="Günlük maksimum kişi sayısı")
        with col4:
            yeni_alan_renk = st.color_picker("Renk", value="#FF6B6B", key="yeni_alan_renk")
        
        if st.button("➕ Alan Ekle", key="alan_ekle_btn"):
            if yeni_alan_isim.strip():
                mevcut_isimler = [a["isim"] for a in st.session_state.get("alanlar", [])]
                if yeni_alan_isim.strip() not in mevcut_isimler:
                    st.session_state.setdefault("alanlar", []).append({
                        "isim": yeni_alan_isim.strip(),
                        "kontenjan": yeni_alan_kontenjan,
                        "max_kontenjan": yeni_alan_max,
                        "renk": yeni_alan_renk
                    })
                    st.rerun()
                else:
                    st.error("Bu isimde bir alan zaten var!")
            else:
                st.error("Alan adı boş olamaz!")
        
        st.divider()
        
        # Mevcut alanlar
        st.markdown("### 📋 Mevcut Alanlar")
        
        alanlar = st.session_state.get("alanlar", [])
        
        if not alanlar:
            st.caption("Henüz alan tanımlanmamış.")
        else:
            toplam_kontenjan = sum(a.get("kontenjan", 1) for a in alanlar)
            # max_kontenjan None olabilir, bu yüzden or kullanıyoruz
            toplam_max = sum(
                (a.get("max_kontenjan") or (a.get("kontenjan", 1) + 2)) 
                for a in alanlar
            )
            st.caption(f"Toplam günlük: Hedef **{toplam_kontenjan}** / Max **{toplam_max}** kişi")
            
            # Başlık satırı
            hcol1, hcol2, hcol3, hcol4 = st.columns([3, 1, 1, 1])
            with hcol2:
                st.caption("Hedef")
            with hcol3:
                st.caption("Max")
            
            for i, alan in enumerate(alanlar):
                col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
                
                with col1:
                    st.markdown(f"<span style='color:{alan.get('renk', '#808080')}'>●</span> **{alan['isim']}**", unsafe_allow_html=True)
                with col2:
                    # Hedef kontenjan değiştirme
                    yeni_kont = st.number_input(
                        "Hedef", 
                        min_value=1, max_value=10, 
                        value=alan.get("kontenjan", 1),
                        key=f"alan_kont_{i}",
                        label_visibility="collapsed"
                    )
                    if yeni_kont != alan.get("kontenjan", 1):
                        st.session_state["alanlar"][i]["kontenjan"] = yeni_kont
                with col3:
                    # Max kontenjan değiştirme
                    current_max = alan.get("max_kontenjan", alan.get("kontenjan", 1) + 2)
                    yeni_max = st.number_input(
                        "Max", 
                        min_value=yeni_kont, max_value=15, 
                        value=current_max,
                        key=f"alan_max_{i}",
                        label_visibility="collapsed"
                    )
                    if yeni_max != current_max:
                        st.session_state["alanlar"][i]["max_kontenjan"] = yeni_max
                with col4:
                    if st.button("🗑️", key=f"alan_sil_{i}"):
                        st.session_state["alanlar"].pop(i)
                        st.rerun()
        
        st.divider()
        
        # Alan bazlı denklik ayarı
        st.markdown("### ⚖️ Denklik Ayarları")
        alan_denklik = st.checkbox(
            "Alan bazlı denklik sağla",
            value=st.session_state.get("alan_bazli_denklik", True),
            help="Her kişi her alandan benzer sayıda nöbet tutar"
        )
        st.session_state["alan_bazli_denklik"] = alan_denklik
        
        st.divider()
        
        # Personel alan yetkinlikleri
        st.markdown("### 👤 Personel Alan Yetkinlikleri")
        st.caption("Boş bırakılan personeller tüm alanlarda çalışabilir.")
        
        personeller = st.session_state.get("personel_list", [])
        alan_isimleri = [a["isim"] for a in alanlar]
        
        if personeller and alan_isimleri:
            for p in personeller:
                mevcut_yetkinlikler = st.session_state.get("personel_alan_yetkinlikleri", {}).get(p, [])
                
                secilen = st.multiselect(
                    f"{p}",
                    options=alan_isimleri,
                    default=[y for y in mevcut_yetkinlikler if y in alan_isimleri],
                    key=f"yetkinlik_{p}",
                    placeholder="Tüm alanlar"
                )
                
                if secilen:
                    st.session_state.setdefault("personel_alan_yetkinlikleri", {})[p] = secilen
                elif p in st.session_state.get("personel_alan_yetkinlikleri", {}):
                    st.session_state["personel_alan_yetkinlikleri"].pop(p, None)
        
        # ====== KIDEM KURALLARI ======
        st.divider()
        st.markdown("### 🎖️ Alan-Kıdem Kuralları")
        st.caption("Her alan için kıdem gruplarından günde min/max kaç kişi olacağını belirleyin.")
        
        kidem_gruplari = st.session_state.get("kidem_gruplari", [])
        grup_isimleri = [g["isim"] for g in kidem_gruplari]
        
        if not kidem_gruplari:
            st.info("Önce Kıdem sekmesinde kıdem grupları tanımlayın.")
        elif alanlar:
            for i, alan in enumerate(alanlar):
                with st.expander(f"📍 {alan['isim']} - Kıdem Kuralları", expanded=False):
                    mevcut_kurallar = alan.get("kidem_kurallari", {})
                    
                    for grup in grup_isimleri:
                        cols = st.columns([3, 1, 1])
                        
                        with cols[0]:
                            st.markdown(f"**{grup}**")
                        
                        with cols[1]:
                            mevcut_min = mevcut_kurallar.get(grup, {}).get("min", 0)
                            yeni_min = st.number_input(
                                f"Min",
                                min_value=0, max_value=10,
                                value=mevcut_min,
                                key=f"kidem_min_{i}_{grup}",
                                help=f"Günde en az kaç {grup}"
                            )
                        
                        with cols[2]:
                            mevcut_max = mevcut_kurallar.get(grup, {}).get("max", 0)
                            yeni_max = st.number_input(
                                f"Max",
                                min_value=0, max_value=10,
                                value=mevcut_max,
                                key=f"kidem_max_{i}_{grup}",
                                help=f"Günde en fazla kaç {grup} (0=sınırsız)"
                            )
                        
                        # Kuralları güncelle
                        if yeni_min > 0 or yeni_max > 0:
                            st.session_state["alanlar"][i].setdefault("kidem_kurallari", {})[grup] = {
                                "min": yeni_min,
                                "max": yeni_max if yeni_max > 0 else None
                            }
                        elif grup in st.session_state["alanlar"][i].get("kidem_kurallari", {}):
                            del st.session_state["alanlar"][i]["kidem_kurallari"][grup]
    else:
        # Alan modu kapalı - bilgi göster
        st.caption("Çoklu alan modu kapalı. Tek alan (eski mod) kullanılacak.")


# =============================================================================
# TAB 3: VARDİYALAR
# =============================================================================

with tabs[3]:
    st.subheader("⏰ Vardiya Tipleri")
    
    st.info("""
    **Vardiya Sistemi**: Farklı süreli vardiyalar tanımlayabilirsiniz (8s, 12s, 24s vs.).
    Her alan için hangi vardiyaların geçerli olduğunu belirleyebilirsiniz.
    """)
    
    # ====== HAZIR ŞABLONLAR ======
    st.markdown("### 📋 Hazır Şablonlar")
    st.caption("Sık kullanılan vardiya tiplerini ekleyin:")
    
    mevcut_vardiyalar = st.session_state.get("vardiya_tipleri", [])
    mevcut_isimler = [v["isim"] for v in mevcut_vardiyalar]
    
    # Şablonları grid olarak göster
    cols = st.columns(4)
    for i, sablon in enumerate(HAZIR_VARDIYALAR):
        with cols[i % 4]:
            zaten_var = sablon.isim in mevcut_isimler
            if st.button(
                f"{'✓ ' if zaten_var else '+'} {sablon.isim}",
                key=f"sablon_{i}",
                disabled=zaten_var,
                use_container_width=True
            ):
                st.session_state.setdefault("vardiya_tipleri", []).append({
                    "isim": sablon.isim,
                    "baslangic": sablon.baslangic,
                    "bitis": sablon.bitis,
                    "renk": sablon.renk
                })
                st.rerun()
    
    st.divider()
    
    # ====== ÖZEL VARDİYA EKLE ======
    st.markdown("### ➕ Özel Vardiya Ekle")
    
    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
    
    with col1:
        yeni_vardiya_isim = st.text_input("Vardiya adı", placeholder="Örn: Özel Gece", key="yeni_vardiya_isim")
    with col2:
        yeni_baslangic = st.time_input("Başlangıç", value=None, key="yeni_vardiya_bas")
    with col3:
        yeni_bitis = st.time_input("Bitiş", value=None, key="yeni_vardiya_bit")
    with col4:
        yeni_vardiya_renk = st.color_picker("Renk", value="#2196F3", key="yeni_vardiya_renk")
    
    if st.button("➕ Vardiya Ekle", key="vardiya_ekle_btn"):
        if yeni_vardiya_isim.strip() and yeni_baslangic and yeni_bitis:
            if yeni_vardiya_isim.strip() not in mevcut_isimler:
                st.session_state.setdefault("vardiya_tipleri", []).append({
                    "isim": yeni_vardiya_isim.strip(),
                    "baslangic": yeni_baslangic.strftime("%H:%M"),
                    "bitis": yeni_bitis.strftime("%H:%M"),
                    "renk": yeni_vardiya_renk
                })
                st.rerun()
            else:
                st.error("Bu isimde bir vardiya zaten var!")
        else:
            st.error("Tüm alanları doldurun!")
    
    st.divider()
    
    # ====== MEVCUT VARDİYALAR ======
    st.markdown("### 📋 Aktif Vardiyalar")
    
    if not mevcut_vardiyalar:
        st.caption("Henüz vardiya tanımlanmamış. Yukarıdan şablon seçin veya özel vardiya ekleyin.")
    else:
        for i, v in enumerate(mevcut_vardiyalar):
            # Saat hesapla
            try:
                vt = VardiyaTipi(v["isim"], v["baslangic"], v["bitis"], v.get("renk", "#808080"))
                saat = vt.saat
            except:
                saat = "?"
            
            col1, col2, col3, col4 = st.columns([4, 2, 1, 1])
            
            with col1:
                st.markdown(
                    f"<span style='color:{v.get('renk', '#808080')}'>●</span> **{v['isim']}**",
                    unsafe_allow_html=True
                )
            with col2:
                st.caption(f"{v['baslangic']} → {v['bitis']} ({saat}s)")
            with col3:
                yeni_renk = st.color_picker(
                    "Renk",
                    value=v.get("renk", "#808080"),
                    key=f"vardiya_renk_{i}",
                    label_visibility="collapsed"
                )
                if yeni_renk != v.get("renk"):
                    st.session_state["vardiya_tipleri"][i]["renk"] = yeni_renk
            with col4:
                if st.button("🗑️", key=f"vardiya_sil_{i}"):
                    st.session_state["vardiya_tipleri"].pop(i)
                    st.rerun()
    
    st.divider()
    
    # ====== ALAN-VARDİYA EŞLEŞTİRME ======
    st.markdown("### 🏢 Alan-Vardiya Eşleştirmesi")
    st.caption("Her alan için hangi vardiyaların geçerli olduğunu seçin.")
    
    alanlar = st.session_state.get("alanlar", [])
    vardiya_isimleri = [v["isim"] for v in mevcut_vardiyalar]
    
    if not alanlar:
        st.info("Önce Alanlar sekmesinde çoklu alan modunu aktifleştirin ve alan ekleyin.")
    elif not mevcut_vardiyalar:
        st.info("Önce yukarıdan vardiya tipleri ekleyin.")
    else:
        for i, alan in enumerate(alanlar):
            mevcut_vardiya_atamalari = alan.get("vardiya_tipleri", [])
            
            secilen = st.multiselect(
                f"📍 {alan['isim']}",
                options=vardiya_isimleri,
                default=[v for v in mevcut_vardiya_atamalari if v in vardiya_isimleri],
                key=f"alan_vardiya_{i}",
                placeholder="Tüm vardiyalar"
            )
            
            st.session_state["alanlar"][i]["vardiya_tipleri"] = secilen
    
    st.divider()
    
    # ====== PERSONEL VARDİYA KISITLARI ======
    st.markdown("### 👤 Personel Vardiya Kısıtları")
    st.caption("Bazı personeller sadece belirli vardiyalarda çalışabilir.")
    
    personeller = st.session_state.get("personel_list", [])
    
    if not personeller:
        st.info("Önce Kişiler sekmesinde personel ekleyin.")
    elif not mevcut_vardiyalar:
        st.info("Önce yukarıdan vardiya tipleri ekleyin.")
    else:
        personel_vardiya_kisitlari = st.session_state.get("personel_vardiya_kisitlari", {})
        
        with st.expander("Vardiya kısıtları düzenle", expanded=False):
            for p in personeller:
                mevcut = personel_vardiya_kisitlari.get(p, [])
                
                secilen = st.multiselect(
                    f"{p}",
                    options=vardiya_isimleri,
                    default=[v for v in mevcut if v in vardiya_isimleri],
                    key=f"personel_vardiya_{p}",
                    placeholder="Tüm vardiyalar"
                )
                
                if secilen:
                    st.session_state.setdefault("personel_vardiya_kisitlari", {})[p] = secilen
                elif p in st.session_state.get("personel_vardiya_kisitlari", {}):
                    del st.session_state["personel_vardiya_kisitlari"][p]
    
    st.divider()
    
    # ====== SAAT DENGESİ AYARI ======
    st.markdown("### ⚖️ Denge Ayarları")
    
    saat_denge = st.checkbox(
        "Saat bazlı denge",
        value=st.session_state.get("saat_bazli_denge", True),
        help="Açık: Toplam çalışma saati dengeli dağıtılır. Kapalı: Vardiya sayısı dengeli dağıtılır."
    )
    st.session_state["saat_bazli_denge"] = saat_denge


# =============================================================================
# TAB 4: İZİNLER
# =============================================================================

with tabs[4]:
    st.subheader("🏖️ İzinler ve Tercihler")
    
    personeller = st.session_state.get("personel_list", [])
    
    if not personeller:
        st.warning("Önce Kişiler sekmesinde personel listesini girin.")
    else:
        yil = int(st.session_state["yil"])
        ay = int(st.session_state["ay"])
        gun_sayisi = ay_gun_sayisi(yil, ay)
        gun_listesi = list(range(1, gun_sayisi + 1))
        
        # İzin map'i hazırla
        izin_map = st.session_state.get("izin_map", {})
        izin_map = {k: v for k, v in izin_map.items() if k in personeller}
        for p in personeller:
            izin_map.setdefault(p, [])
        st.session_state["izin_map"] = izin_map
        
        # Her personel için izin girişi
        for p in personeller:
            with st.expander(f"📅 {p}", expanded=False):
                selected = st.multiselect(
                    "İzinli günler",
                    options=gun_listesi,
                    default=sorted(list(set(st.session_state["izin_map"].get(p, [])))),
                    key=f"izin_{p}"
                )
                st.session_state["izin_map"][p] = sorted(selected)
                
                # Bloklu hafta günleri
                gun_adlari = tum_hafta_gunleri()
                st.session_state["weekday_block_map"].setdefault(p, [])
                # Sadece geçerli gün adlarını default olarak al
                mevcut_bloklar = st.session_state["weekday_block_map"].get(p, [])
                valid_bloklar = [g for g in mevcut_bloklar if g in gun_adlari]
                blocked = st.multiselect(
                    "Bloklu hafta günleri (her hafta)",
                    options=gun_adlari,
                    default=valid_bloklar,
                    key=f"wblock_{p}"
                )
                st.session_state["weekday_block_map"][p] = blocked
                
                # Tercih edilen günler
                st.session_state.setdefault("prefer_map", {}).setdefault(p, [])
                prefer_selected = st.multiselect(
                    "Tercih edilen günler (soft)",
                    options=gun_listesi,
                    default=sorted(list(set(st.session_state["prefer_map"].get(p, [])))),
                    key=f"prefer_{p}"
                )
                st.session_state["prefer_map"][p] = sorted(prefer_selected)
        
        st.divider()
        toplam_izin = sum(len(v) for v in st.session_state["izin_map"].values())
        st.caption(f"✓ Toplam izin günü: {toplam_izin}")
        
        # Tatiller
        st.divider()
        st.subheader("🎌 Resmi Tatiller")
        
        auto_holidays = resmi_tatiller(yil, ay)
        
        if auto_holidays:
            st.success("✓ Bu ay için otomatik tespit edilen tatiller:")
            for gun, isim in sorted(auto_holidays.items()):
                st.write(f"  • {gun}. gün - {isim}")
        else:
            st.info("Bu ay resmi tatil bulunmuyor.")
        
        manuel_input = st.text_input(
            "Ekstra tatil günleri (örn: 15, 16)",
            value=st.session_state.get("manuel_tatiller", ""),
            key="manuel_tatiller_input"
        )
        st.session_state["manuel_tatiller"] = manuel_input
        
        if manuel_input.strip():
            manuel_gunler = gun_parse(manuel_input, gun_sayisi)
            if manuel_gunler:
                st.caption(f"  → Eklenecek: {sorted(manuel_gunler)}")


# =============================================================================
# TAB 5: EŞLEŞMELER
# =============================================================================

with tabs[5]:
    st.subheader("👫 Eşleşme Tercihleri")
    
    personeller = st.session_state.get("personel_list", [])
    
    if not personeller:
        st.warning("Önce Kişiler sekmesinde personel listesini girin.")
    elif len(personeller) < 2:
        st.warning("Çift tanımlamak için en az 2 personel gerekli.")
    else:
        colA, colB = st.columns(2)
        
        with colA:
            st.markdown("### ✅ Birlikte Tutsun")
            a = st.selectbox("Personel A", options=personeller, key="wp_a")
            b_options = [p for p in personeller if p != a]
            b = st.selectbox("Personel B", options=b_options, key="wp_b")
            min_k = st.number_input("Minimum birlikte gün", min_value=1, max_value=31, value=2, key="wp_min")
            
            if st.button("➕ Ekle", key="wp_add"):
                aa, bb = sorted([a, b])
                exists = any(
                    item["a"] == aa and item["b"] == bb 
                    for item in st.session_state["want_pairs_list"]
                )
                if not exists:
                    st.session_state["want_pairs_list"].append({"a": aa, "b": bb, "min": int(min_k)})
                    st.rerun()
        
        with colB:
            st.markdown("### ❌ Asla Birlikte Tutmasın")
            na = st.selectbox("Personel A ", options=personeller, key="np_a")
            nb_options = [p for p in personeller if p != na]
            nb = st.selectbox("Personel B ", options=nb_options, key="np_b")
            
            if st.button("➕ Ekle", key="np_add"):
                aa, bb = sorted([na, nb])
                exists = any(
                    item["a"] == aa and item["b"] == bb 
                    for item in st.session_state["no_pairs_list"]
                )
                if not exists:
                    st.session_state["no_pairs_list"].append({"a": aa, "b": bb})
                    st.rerun()
        
        st.divider()
        st.markdown("### Mevcut Tanımlar")
        
        colL, colR = st.columns(2)
        
        with colL:
            st.markdown("**Birlikte tutulacaklar:**")
            if not st.session_state["want_pairs_list"]:
                st.caption("Henüz yok.")
            else:
                for i, item in enumerate(st.session_state["want_pairs_list"]):
                    c1, c2 = st.columns([6, 2])
                    c1.write(f"• {item['a']} ↔ {item['b']} (min: {item['min']})")
                    if c2.button("Sil", key=f"wp_del_{i}"):
                        st.session_state["want_pairs_list"].pop(i)
                        st.rerun()
        
        with colR:
            st.markdown("**Ayrı tutulacaklar:**")
            if not st.session_state["no_pairs_list"]:
                st.caption("Henüz yok.")
            else:
                for i, item in enumerate(st.session_state["no_pairs_list"]):
                    c1, c2 = st.columns([6, 2])
                    c1.write(f"• {item['a']} × {item['b']}")
                    if c2.button("Sil", key=f"np_del_{i}"):
                        st.session_state["no_pairs_list"].pop(i)
                        st.rerun()
        
        # Gelişmiş ayarlar
        with st.expander("⚙️ Gelişmiş Ayarlar"):
            st.markdown("#### ☁️ Esnek Ayrı Tutma (Soft)")
            sna = st.selectbox("Personel A", options=personeller, key="snp_a")
            snb_options = [p for p in personeller if p != sna]
            snb = st.selectbox("Personel B", options=snb_options, key="snp_b")
            
            if st.button("➕ Esnek kural ekle"):
                aa, bb = sorted([sna, snb])
                exists = any(
                    item["a"] == aa and item["b"] == bb 
                    for item in st.session_state["soft_no_pairs_list"]
                )
                if not exists:
                    st.session_state["soft_no_pairs_list"].append({"a": aa, "b": bb})
                    st.rerun()
            
            for i, item in enumerate(st.session_state["soft_no_pairs_list"]):
                c1, c2 = st.columns([8, 2])
                c1.write(f"☁️ {item['a']} - {item['b']}")
                if c2.button("Sil", key=f"snp_del_{i}"):
                    st.session_state["soft_no_pairs_list"].pop(i)
                    st.rerun()
            
            st.divider()
            
            # === KURAL AYARLARI ===
            st.markdown("#### 📋 Nöbet Kuralları")
            
            # Ardışık gün yasağı
            st.session_state["ardisik_yasak"] = st.checkbox(
                "Ardışık gün yasağı",
                value=st.session_state.get("ardisik_yasak", True),
                help="Aynı kişi arka arkaya iki gün nöbet tutamaz"
            )
            
            # Günaşırı limiti
            col1, col2 = st.columns([1, 2])
            with col1:
                st.session_state["gunasiri_limit_aktif"] = st.checkbox(
                    "Günaşırı limit",
                    value=st.session_state.get("gunasiri_limit_aktif", True),
                    help="1 gün arayla nöbet sayısını sınırla"
                )
            with col2:
                if st.session_state.get("gunasiri_limit_aktif", True):
                    st.session_state["max_gunasiri"] = st.number_input(
                        "Maksimum günaşırı nöbet (kişi başı/ay)",
                        min_value=1, max_value=15, 
                        value=st.session_state.get("max_gunasiri", 1),
                        help="0 = sınırsız"
                    )
            
            st.divider()
            st.markdown("#### ⚖️ Denge Kuralları")
            
            # Hafta sonu dengesi
            st.session_state["hafta_sonu_dengesi"] = st.checkbox(
                "Hafta sonu dengesi",
                value=st.session_state.get("hafta_sonu_dengesi", True),
                help="Cuma, Cumartesi, Pazar nöbetlerini dengeli dağıt"
            )
            
            if st.session_state.get("hafta_sonu_dengesi", True):
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.session_state["w_cuma"] = st.slider("Cuma ağırlığı", 0, 2000, 
                        st.session_state.get("w_cuma", 1000))
                with col2:
                    st.session_state["w_cumartesi"] = st.slider("Cumartesi ağırlığı", 0, 2000,
                        st.session_state.get("w_cumartesi", 1000))
                with col3:
                    st.session_state["w_pazar"] = st.slider("Pazar ağırlığı", 0, 2000,
                        st.session_state.get("w_pazar", 1000))
            
            # Tatil dengesi
            st.session_state["tatil_dengesi"] = st.checkbox(
                "Tatil dengesi",
                value=st.session_state.get("tatil_dengesi", True),
                help="Resmi tatil nöbetlerini dengeli dağıt"
            )
            
            st.divider()
            st.markdown("#### 🎯 Tercihler")
            
            # 2 gün boşluk tercihi
            col1, col2 = st.columns([1, 2])
            with col1:
                st.session_state["iki_gun_bosluk_aktif"] = st.checkbox(
                    "2 gün boşluk tercihi",
                    value=st.session_state.get("iki_gun_bosluk_aktif", True),
                    help="Nöbetler arası en az 2 gün boşluk tercih edilir"
                )
            with col2:
                if st.session_state.get("iki_gun_bosluk_aktif", True):
                    st.session_state["w_gap3"] = st.slider(
                        "Tercih ağırlığı",
                        0, 2000, st.session_state.get("w_gap3", 300)
                    )


# =============================================================================
# TAB 6: ÇÖZÜM
# =============================================================================

with tabs[6]:
    st.subheader("✅ Çözüm")
    
    # Demo modunda özet göster
    if is_demo_active():
        meta = get_demo_meta()
        st.success(f"🧪 Demo senaryosu hazır! Zorluk: **{meta.get('difficulty')}** | Seed: `{meta.get('seed')}`")
    
    if st.button("🚀 Nöbeti Oluştur", type="primary", use_container_width=True):
        yil = int(st.session_state["yil"])
        ay = int(st.session_state["ay"])
        default_target = int(st.session_state.get("varsayilan_hedef", 7))
        personeller = st.session_state.get("personel_list", [])
        
        if not personeller:
            st.error("Personel listesi boş olamaz.")
            st.stop()
        
        gun_sayisi = ay_gun_sayisi(yil, ay)
        
        # Hedefler - öncelik: kişisel > kıdem grubu > genel varsayılan
        hedefler = {}  # Toplam nöbet hedefi
        vardiya_hedefleri = {}  # {kisi: {vardiya: hedef}} - vardiya bazlı hedefler
        
        personel_kidem = st.session_state.get("personel_kidem_gruplari", {})
        kidem_gruplari = st.session_state.get("kidem_gruplari", [])
        vardiyalar_data = st.session_state.get("vardiya_tipleri", [])
        
        # Kıdem grubu hedeflerini dict'e çevir
        grup_hedefleri = {
            g["isim"]: g.get("varsayilan_hedef", default_target)
            for g in kidem_gruplari
        }
        
        # Kıdem grubu vardiya hedeflerini dict'e çevir
        grup_vardiya_hedefleri = {
            g["isim"]: g.get("vardiya_hedefleri", {})
            for g in kidem_gruplari
        }
        
        for p in personeller:
            # Önce kişisel hedefe bak
            kisisel_hedef = st.session_state.get("personel_targets", {}).get(p)
            kidem = personel_kidem.get(p)
            
            if kisisel_hedef is not None and kisisel_hedef != default_target:
                # Kişisel hedef var
                hedefler[p] = kisisel_hedef
            elif kidem and kidem in grup_hedefleri:
                # Kıdem grubunun hedefine bak
                hedefler[p] = grup_hedefleri[kidem]
                
                # Vardiya bazlı hedef var mı?
                if vardiyalar_data and kidem in grup_vardiya_hedefleri:
                    v_hedef = grup_vardiya_hedefleri[kidem]
                    if v_hedef and any(v > 0 for v in v_hedef.values()):
                        vardiya_hedefleri[p] = v_hedef
            else:
                # Genel varsayılan
                hedefler[p] = default_target
        
        # İzinler (set olarak)
        izinler = {}
        for p, gunler in st.session_state.get("izin_map", {}).items():
            izinler[p] = set(gunler) if gunler else set()
        
        # Hafta günü bloklarını izinlere ekle
        for p in personeller:
            blocked_names = st.session_state.get("weekday_block_map", {}).get(p, [])
            for gun_adi in blocked_names:
                wd = hafta_gunu_numarasi(gun_adi)
                if wd >= 0:
                    for gun in range(1, gun_sayisi + 1):
                        if datetime(yil, ay, gun).weekday() == wd:
                            izinler.setdefault(p, set()).add(gun)
        
        # Tercih edilen günler
        tercih_edilen = {}
        for p, gunler in st.session_state.get("prefer_map", {}).items():
            tercih_edilen[p] = set(gunler) if gunler else set()
        
        # Tatiller
        auto_holidays = set(resmi_tatiller(yil, ay).keys())
        manuel_text = st.session_state.get("manuel_tatiller", "")
        manuel_holidays = gun_parse(manuel_text, gun_sayisi) if manuel_text.strip() else set()
        tatiller = auto_holidays | manuel_holidays
        
        # Eşleşme kuralları
        ayri_tut = [
            (item["a"], item["b"]) 
            for item in st.session_state.get("no_pairs_list", [])
        ]
        birlikte_tut = [
            (item["a"], item["b"], int(item["min"])) 
            for item in st.session_state.get("want_pairs_list", [])
        ]
        esnek_ayri_tut = [
            (item["a"], item["b"]) 
            for item in st.session_state.get("soft_no_pairs_list", [])
        ]
        
        # Toplam hedef hesapla (feasibility kontrolü için)
        toplam_hedef = sum(hedefler.values())
        
        # Çoklu alan modu kontrolü
        alan_modu_aktif = st.session_state.get("alan_modu_aktif", False)
        alanlar_data = st.session_state.get("alanlar", [])
        
        if alan_modu_aktif and alanlar_data:
            alanlar = [
                AlanTanimi(
                    isim=a["isim"], 
                    gunluk_kontenjan=a.get("kontenjan", 1),
                    max_kontenjan=a.get("max_kontenjan"),
                    kidem_kurallari=a.get("kidem_kurallari", {}),
                    vardiya_tipleri=a.get("vardiya_tipleri", [])
                )
                for a in alanlar_data
            ]
            toplam_kontenjan = sum(a.gunluk_kontenjan for a in alanlar)
            gereken_toplam = toplam_kontenjan * gun_sayisi
            
            if toplam_hedef < gereken_toplam:
                st.error(f"İmkânsız: Toplam hedef ({toplam_hedef}) < gereken ({gereken_toplam} = {toplam_kontenjan}/gün x {gun_sayisi} gün)")
                st.stop()
        else:
            alanlar = []
            if toplam_hedef < gun_sayisi:
                st.error(f"İmkânsız: Toplam hedef ({toplam_hedef}) < gün sayısı ({gun_sayisi})")
                st.stop()
        
        # Vardiya tipleri
        vardiyalar_data = st.session_state.get("vardiya_tipleri", [])
        vardiyalar = [
            VardiyaTanimi(
                isim=v["isim"],
                baslangic=v.get("baslangic", "08:00"),
                bitis=v.get("bitis", "16:00")
            )
            for v in vardiyalar_data
        ]
        
        # Personel alan yetkinlikleri
        personel_alan_yetkinlikleri = st.session_state.get("personel_alan_yetkinlikleri", {})
        
        # Personel vardiya kısıtları
        personel_vardiya_kisitlari = st.session_state.get("personel_vardiya_kisitlari", {})
        
        # Solver config - kullanıcı ayarlarından al
        config = SolverConfig(
            # Hard constraints
            ardisik_yasak=st.session_state.get("ardisik_yasak", True),
            gunasiri_limit_aktif=st.session_state.get("gunasiri_limit_aktif", True),
            max_gunasiri_per_kisi=st.session_state.get("max_gunasiri", 1),
            
            # Hafta sonu dengesi
            hafta_sonu_dengesi_aktif=st.session_state.get("hafta_sonu_dengesi", True),
            w_cuma=st.session_state.get("w_cuma", 1000),
            w_cumartesi=st.session_state.get("w_cumartesi", 1000),
            w_pazar=st.session_state.get("w_pazar", 1000),
            
            # Tatil dengesi
            tatil_dengesi_aktif=st.session_state.get("tatil_dengesi", True),
            
            # 2 gün boşluk tercihi
            iki_gun_bosluk_aktif=st.session_state.get("iki_gun_bosluk_aktif", True),
            w_iki_gun_bosluk=st.session_state.get("w_gap3", 300),
            
            # Saat bazlı denge
            saat_bazli_denge=st.session_state.get("saat_bazli_denge", True)
        )
        
        # Solver input
        solver_input = SolverInput(
            yil=yil,
            ay=ay,
            personeller=personeller,
            hedefler=hedefler,
            vardiya_hedefleri=vardiya_hedefleri,
            izinler=izinler,
            tatiller=tatiller,
            ayri_tut=ayri_tut,
            birlikte_tut=birlikte_tut,
            esnek_ayri_tut=esnek_ayri_tut,
            tercih_edilen=tercih_edilen,
            alanlar=alanlar,
            personel_alan_yetkinlikleri=personel_alan_yetkinlikleri,
            alan_bazli_denklik=st.session_state.get("alan_bazli_denklik", True),
            personel_kidem_gruplari=st.session_state.get("personel_kidem_gruplari", {}),
            vardiyalar=vardiyalar,
            personel_vardiya_kisitlari=personel_vardiya_kisitlari,
            config=config
        )
        
        mod_bilgi = []
        if alanlar:
            mod_bilgi.append("Çoklu alan")
        if vardiyalar:
            mod_bilgi.append("Vardiya")
        if vardiya_hedefleri:
            mod_bilgi.append("Vardiya hedefleri")
        mod_str = f" ({', '.join(mod_bilgi)})" if mod_bilgi else ""
        st.info(f"Solver çalıştırılıyor...{mod_str}")
        
        try:
            solver = NobetSolver(solver_input)
            schedule = solver.coz()
            
            # Planı kaydet
            plan = AylikPlan(
                yil=yil,
                ay=ay,
                izinler={p: list(g) for p, g in izinler.items()},
                tercih_edilen_gunler={p: list(g) for p, g in tercih_edilen.items()},
                manuel_tatiller=list(manuel_holidays),
                hedef_override={p: h for p, h in hedefler.items() if h != default_target},
                sonuc={str(k): v for k, v in schedule.items()},
                sonuc_alanlı=bool(alanlar)
            )
            aylik_plani_kaydet(plan)
            
        except Exception as e:
            st.error("❌ Çözüm bulunamadı.")
            st.caption(str(e))
            
            # Gelişmiş teşhis
            from solver import gelismis_teshis, teshis_ozeti, TeshisSonucu
            
            teshisler = gelismis_teshis(
                yil=yil,
                ay=ay,
                personeller=personeller,
                hedefler=hedefler,
                vardiya_hedefleri=vardiya_hedefleri,
                izinler=izinler,
                tatiller=tatiller,
                birlikte_tut=birlikte_tut,
                ayri_tut=ayri_tut,
                alanlar=alanlar if alanlar else None,
                vardiyalar=vardiyalar if vardiyalar else None,
                personel_alan_yetkinlikleri=personel_alan_yetkinlikleri,
                personel_vardiya_kisitlari=personel_vardiya_kisitlari,
                personel_kidem_gruplari=st.session_state.get("personel_kidem_gruplari", {}),
                ardisik_yasak=st.session_state.get("ardisik_yasak", True)
            )
            
            st.warning("🔍 **Tespit Edilen Sorunlar:**")
            
            errors = [t for t in teshisler if t.seviye == "error"]
            warnings = [t for t in teshisler if t.seviye == "warning"]
            
            if errors:
                st.markdown(f"**❌ {len(errors)} Kritik Sorun:**")
                for t in errors[:10]:
                    with st.expander(f"🔴 {t.mesaj}", expanded=True):
                        st.json(t.detay)
            
            if warnings:
                st.markdown(f"**⚠️ {len(warnings)} Uyarı:**")
                for t in warnings[:5]:
                    with st.expander(f"🟡 {t.mesaj}", expanded=False):
                        st.json(t.detay)
            
            st.stop()
        
        # Sonuç tablosu - mod'a göre farklı gösterim
        weekdays_tr = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        
        # Mod tespiti
        has_alanlar = bool(alanlar)
        has_vardiyalar = bool(vardiyalar)
        
        if has_alanlar and has_vardiyalar:
            # ALAN + VARDİYA MODU - {gun: {alan: {vardiya: [kişiler]}}}
            alan_isimleri = [a.isim for a in alanlar]
            vardiya_isimleri = [v.isim for v in vardiyalar]
            
            rows = []
            for gun in range(1, gun_sayisi + 1):
                dt = datetime(yil, ay, gun)
                wd = weekdays_tr[dt.weekday()]
                gun_data = schedule.get(gun, {})
                
                row = {
                    "Gün": gun,
                    "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                    "Hafta Günü": wd,
                    "Tatil": "Evet" if gun in tatiller else "",
                }
                
                # Her alan-vardiya kombinasyonu için sütun
                for alan_isim in alan_isimleri:
                    alan_data = gun_data.get(alan_isim, {})
                    for vardiya_isim in vardiya_isimleri:
                        kisiler = alan_data.get(vardiya_isim, [])
                        col_name = f"{alan_isim} / {vardiya_isim}"
                        row[col_name] = ", ".join(kisiler) if kisiler else "-"
                
                rows.append(row)
            
            df_schedule = pd.DataFrame(rows)
            
            st.success("🎉 Çözüm bulundu! (Çoklu Alan + Vardiya)")
            st.subheader("📋 Oluşturulan Nöbet Listesi")
            st.dataframe(df_schedule, use_container_width=True, hide_index=True)
            
            # İstatistikler
            st.divider()
            st.subheader("📊 Personel Dağılımı")
            
            stats = []
            for p in personeller:
                stat = {"Personel": p}
                toplam = 0
                toplam_saat = 0
                for gun_data in schedule.values():
                    for alan_isim, alan_data in gun_data.items():
                        if isinstance(alan_data, dict):
                            for vardiya_isim, kisiler in alan_data.items():
                                if p in kisiler:
                                    toplam += 1
                                    # Saat hesapla
                                    for v in vardiyalar:
                                        if v.isim == vardiya_isim:
                                            toplam_saat += v.saat
                                            break
                stat["Toplam Nöbet"] = toplam
                stat["Toplam Saat"] = toplam_saat
                stat["Hedef"] = hedefler.get(p, default_target)
                stats.append(stat)
            
            st.table(pd.DataFrame(stats))
        
        elif has_vardiyalar:
            # SADECE VARDİYA MODU - {gun: {vardiya: [kişiler]}}
            vardiya_isimleri = [v.isim for v in vardiyalar]
            
            rows = []
            for gun in range(1, gun_sayisi + 1):
                dt = datetime(yil, ay, gun)
                wd = weekdays_tr[dt.weekday()]
                gun_data = schedule.get(gun, {})
                
                row = {
                    "Gün": gun,
                    "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                    "Hafta Günü": wd,
                    "Tatil": "Evet" if gun in tatiller else "",
                }
                
                # Her vardiya için sütun
                for vardiya_isim in vardiya_isimleri:
                    kisiler = gun_data.get(vardiya_isim, [])
                    row[vardiya_isim] = ", ".join(kisiler) if kisiler else "-"
                
                rows.append(row)
            
            df_schedule = pd.DataFrame(rows)
            
            st.success("🎉 Çözüm bulundu! (Vardiya Modu)")
            st.subheader("📋 Oluşturulan Nöbet Listesi")
            st.dataframe(df_schedule, use_container_width=True, hide_index=True)
            
            # Vardiya bazlı dağılım istatistikleri
            st.divider()
            st.subheader("📊 Vardiya Bazlı Dağılım")
            
            stats = []
            for p in personeller:
                stat = {"Personel": p}
                toplam = 0
                toplam_saat = 0
                for vardiya in vardiyalar:
                    count = sum(1 for g_data in schedule.values() if p in g_data.get(vardiya.isim, []))
                    stat[vardiya.isim] = count
                    toplam += count
                    toplam_saat += count * vardiya.saat
                stat["TOPLAM"] = toplam
                stat["Saat"] = toplam_saat
                stat["Hedef"] = hedefler.get(p, default_target)
                stats.append(stat)
            
            st.table(pd.DataFrame(stats))
        
        elif has_alanlar:
            # ÇOKLU ALAN MODU - sonuç formatı: {gun: {alan: [kişiler]}}
            alan_isimleri = [a.isim for a in alanlar]
            
            rows = []
            for gun in range(1, gun_sayisi + 1):
                dt = datetime(yil, ay, gun)
                wd = weekdays_tr[dt.weekday()]
                gun_data = schedule.get(gun, {})
                
                row = {
                    "Gün": gun,
                    "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                    "Hafta Günü": wd,
                    "Tatil": "Evet" if gun in tatiller else "",
                }
                
                # Her alan için sütun
                for alan_isim in alan_isimleri:
                    kisiler = gun_data.get(alan_isim, [])
                    row[alan_isim] = ", ".join(kisiler) if kisiler else "-"
                
                rows.append(row)
            
            df_schedule = pd.DataFrame(rows)
            
            st.success("🎉 Çözüm bulundu! (Çoklu Alan Modu)")
            st.subheader("📋 Oluşturulan Nöbet Listesi")
            st.dataframe(df_schedule, use_container_width=True, hide_index=True)
            
            # Alan bazlı dağılım istatistikleri
            st.divider()
            st.subheader("📊 Alan Bazlı Dağılım")
            
            alan_stats = []
            for p in personeller:
                stat = {"Personel": p}
                toplam = 0
                for alan_isim in alan_isimleri:
                    count = sum(1 for g in schedule.values() if p in g.get(alan_isim, []))
                    stat[alan_isim] = count
                    toplam += count
                stat["TOPLAM"] = toplam
                stat["Hedef"] = hedefler.get(p, default_target)
                alan_stats.append(stat)
            
            st.table(pd.DataFrame(alan_stats))
            
        else:
            # TEK ALAN MODU - eski format: {gun: [kişiler]}
            max_kisi = max((len(v) for v in schedule.values() if isinstance(v, list)), default=1)
        
            rows = []
            for gun in range(1, gun_sayisi + 1):
                dt = datetime(yil, ay, gun)
                wd = weekdays_tr[dt.weekday()]
                isimler = schedule.get(gun, [])
                if not isinstance(isimler, list):
                    isimler = []
                row = {
                    "Gün": gun,
                    "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                    "Hafta Günü": wd,
                    "Kişi Sayısı": len(isimler),
                    "Tatil": "Evet" if gun in tatiller else "",
                }
                for i in range(max_kisi):
                    row[f"Nöbetçi {i+1}"] = isimler[i] if i < len(isimler) else ""
                rows.append(row)
            
            df_schedule = pd.DataFrame(rows)
            
            st.success("🎉 Çözüm bulundu!")
            st.subheader("📋 Oluşturulan Nöbet Listesi")
            st.dataframe(df_schedule, use_container_width=True, hide_index=True)
            
            # Personel dağılımı
            st.divider()
            st.subheader("📊 Personel Nöbet Dağılımı")
            stats = []
            for p in personeller:
                count = sum(1 for v in schedule.values() if isinstance(v, list) and p in v)
                hedef = hedefler.get(p, default_target)
                stats.append({
                    "Personel": p,
                    "Hedef": hedef,
                    "Gerçekleşen": count,
                    "Fark": count - hedef
                })
            
            st.table(pd.DataFrame(stats))
        
        # CSV indirme (her iki mod için)
        csv_data = df_schedule.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            "📥 CSV İndir",
            data=csv_data,
            file_name=f"Nobet_{yil}_{ay:02d}.csv",
            mime="text/csv"
        )
        
        # Excel indirme
        xlsx_buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Nöbet {ay:02d}-{yil}"
        
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        
        fieldnames = list(rows[0].keys())
        for c, h in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        fill_weekend = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        fill_holiday = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        
        for r_i, row in enumerate(rows, start=2):
            dt = datetime(yil, ay, row["Gün"])
            is_weekend = weekdays_tr[dt.weekday()] in ["Cuma", "Cumartesi", "Pazar"]
            is_holiday = row["Gün"] in tatiller
            
            for c_i, h in enumerate(fieldnames, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=row.get(h, ""))
                if c_i <= 5:
                    cell.alignment = center
                
                if is_holiday:
                    cell.fill = fill_holiday
                elif is_weekend:
                    cell.fill = fill_weekend
        
        for col in ws.columns:
            maxlen = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 30)
        
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)
        
        st.download_button(
            "⬇️ Excel İndir (XLSX)",
            data=xlsx_buf.getvalue(),
            file_name=f"nobet_{ay:02d}_{yil}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
