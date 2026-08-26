"""
Çözüm sekmesi — Solver çalıştırma, sonuç gösterim ve dışa aktarma.
"""

import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import AylikPlan
from utils import (
    ay_gun_sayisi, resmi_tatiller, gun_parse,
    hafta_gunu_adi, tum_hafta_gunleri, hafta_gunu_numarasi,
    hesapla_otomatik_hedef
)
from storage import aylik_plani_kaydet, onceki_ay_son_gun_atamalari
from solver import (
    NobetSolver, SolverInput, SolverConfig,
    AlanTanimi, VardiyaTanimi, gelismis_teshis
)
from streamlit_integration import is_demo_active, get_demo_meta
from config import (
    w_cuma, w_cumartesi, w_pazar, w_iki_gun_bosluk, max_gunasiri_per_kisi,
)


def _staffing_taban_tavan(alanlar, vardiyalar, gun_sayisi, enforce_minimum_staffing):
    """
    Günlük zorunlu doluluk tabanı ve (varsa) teorik doluluk tavanını hesaplar.

    Solver'daki HARD kısıt döngüleriyle (_vardiya_minimum_kontenjan_hard,
    _alan_kontenjan_soft'un max_kontenjan hard-cap'i) BİREBİR AYNI kombinasyon
    filtresini kullanır: alan.vardiya_tipleri tanımlıysa ve bu vardiya o
    listede yoksa kombinasyon sayılmaz.

    Döner: (hard_taban, teorik_tavan). teorik_tavan tanımsızsa (kapasite üst
    sınırı olmayan bir mod veya en az bir alanda max_kontenjan boşsa) None.
    """
    coklu_alan_modu = bool(alanlar)
    vardiya_modu = bool(vardiyalar)

    hard_taban_gunluk = 0
    if vardiya_modu and enforce_minimum_staffing:
        if coklu_alan_modu:
            for alan in alanlar:
                for vardiya in vardiyalar:
                    if alan.vardiya_tipleri and vardiya.isim not in alan.vardiya_tipleri:
                        continue
                    hard_taban_gunluk += max(alan.minimum_staffing, vardiya.minimum_staffing)
        else:
            for vardiya in vardiyalar:
                hard_taban_gunluk += vardiya.minimum_staffing

    teorik_tavan_gunluk = 0
    tavan_tanimli = coklu_alan_modu
    if coklu_alan_modu:
        vardiya_dongu = vardiyalar if vardiya_modu else [None]
        for alan in alanlar:
            for vardiya in vardiya_dongu:
                if vardiya is not None and vardiya_modu and alan.vardiya_tipleri and vardiya.isim not in alan.vardiya_tipleri:
                    continue
                if not alan.max_kontenjan or alan.max_kontenjan <= 0:
                    tavan_tanimli = False
                    continue
                teorik_tavan_gunluk += alan.max_kontenjan

    hard_taban = hard_taban_gunluk * gun_sayisi
    teorik_tavan = teorik_tavan_gunluk * gun_sayisi if tavan_tanimli else None
    return hard_taban, teorik_tavan


def render_cozum_tab():
    st.subheader("✅ Çözüm")

    # Demo modunda özet göster
    if is_demo_active():
        meta = get_demo_meta()
        st.success(f"🧪 Demo senaryosu hazır! Zorluk: **{meta.get('difficulty')}** | Seed: `{meta.get('seed')}`")

    if st.button("🚀 Nöbeti Oluştur", type="primary", use_container_width=True):
        _cozum_olustur()


def _cozum_olustur():
    yil = int(st.session_state["yil"])
    ay = int(st.session_state["ay"])
    default_target = int(st.session_state.get("varsayilan_hedef", 7))
    personeller = st.session_state.get("personel_list", [])

    if not personeller:
        st.error("Personel listesi boş olamaz.")
        st.stop()

    gun_sayisi = ay_gun_sayisi(yil, ay)

    # Hedefler - öncelik: kişisel > otomatik > kıdem grubu > genel varsayılan
    hedefler = {}  # Toplam nöbet hedefi
    vardiya_hedefleri = {}  # {kisi: {vardiya: hedef}} - vardiya bazlı hedefler

    personel_kidem = st.session_state.get("personel_kidem_gruplari", {})
    kidem_gruplari = st.session_state.get("kidem_gruplari", [])
    vardiyalar_data = st.session_state.get("vardiya_tipleri", [])

    # Kıdem grubu hedeflerini dict'e çevir
    grup_hedefleri = {
        g["isim"]: g.get("varsayilan_hedef", default_target)
        for g in kidem_gruplari
    }

    # Kıdem grubu vardiya hedeflerini dict'e çevir
    grup_vardiya_hedefleri = {
        g["isim"]: g.get("vardiya_hedefleri", {})
        for g in kidem_gruplari
    }

    # Otomatik hedef hesaplama (eğer aktifse)
    otomatik_aktif = st.session_state.get("otomatik_hedef", True)
    otomatik_hedefler = {}
    if otomatik_aktif:
        alanlar_data = st.session_state.get("alanlar", [])
        izin_map = st.session_state.get("izin_map", {})
        ardisik = st.session_state.get("ardisik_yasak", True)
        otomatik_hedefler = hesapla_otomatik_hedef(
            gun_sayisi, alanlar_data, vardiyalar_data, personeller, izin_map, ardisik
        )

    for p in personeller:
        # Önce kişisel hedefe bak
        kisisel_hedef = st.session_state.get("personel_targets", {}).get(p)
        kidem = personel_kidem.get(p)

        if kisisel_hedef is not None:
            # Kişisel hedef var (kullanıcı açıkça girmiş)
            hedefler[p] = kisisel_hedef
        elif otomatik_aktif and p in otomatik_hedefler:
            # Otomatik hesaplanan hedef
            hedefler[p] = otomatik_hedefler[p]
        elif kidem and kidem in grup_hedefleri:
            # Kıdem grubunun hedefine bak
            hedefler[p] = grup_hedefleri[kidem]

            # Vardiya bazlı hedef var mı?
            if vardiyalar_data and kidem in grup_vardiya_hedefleri:
                v_hedef = grup_vardiya_hedefleri[kidem]
                if v_hedef and any(v > 0 for v in v_hedef.values()):
                    vardiya_hedefleri[p] = v_hedef
        else:
            # Genel varsayılan
            hedefler[p] = default_target

    # İzinler (set olarak)
    izinler = {}
    for p, gunler in st.session_state.get("izin_map", {}).items():
        izinler[p] = set(gunler) if gunler else set()

    # Hafta günü bloklarını izinlere ekle
    for p in personeller:
        blocked_names = st.session_state.get("weekday_block_map", {}).get(p, [])
        for gun_adi in blocked_names:
            wd = hafta_gunu_numarasi(gun_adi)
            if wd >= 0:
                for gun in range(1, gun_sayisi + 1):
                    if datetime(yil, ay, gun).weekday() == wd:
                        izinler.setdefault(p, set()).add(gun)

    # Tercih edilen günler
    tercih_edilen = {}
    for p, gunler in st.session_state.get("prefer_map", {}).items():
        tercih_edilen[p] = set(gunler) if gunler else set()

    # Tatiller
    auto_holidays = set(resmi_tatiller(yil, ay).keys())
    manuel_text = st.session_state.get("manuel_tatiller", "")
    manuel_holidays = gun_parse(manuel_text, gun_sayisi) if manuel_text.strip() else set()
    tatiller = auto_holidays | manuel_holidays

    # Eşleşme kuralları
    ayri_tut = [
        (item["a"], item["b"])
        for item in st.session_state.get("no_pairs_list", [])
    ]
    birlikte_tut = [
        (item["a"], item["b"], int(item["min"]))
        for item in st.session_state.get("want_pairs_list", [])
    ]
    esnek_ayri_tut = [
        (item["a"], item["b"])
        for item in st.session_state.get("soft_no_pairs_list", [])
    ]

    # Vardiya tipleri (taban/tavan hesabı vardiyalara ihtiyaç duyduğu için önce kurulur)
    vardiyalar = [
        VardiyaTanimi(
            isim=v["isim"],
            baslangic=v.get("baslangic", "08:00"),
            bitis=v.get("bitis", "16:00"),
            minimum_staffing=v.get("minimum_staffing", 1)
        )
        for v in vardiyalar_data
    ]

    # Çoklu alan modu kontrolü
    alan_modu_aktif = st.session_state.get("alan_modu_aktif", False)
    alanlar_data = st.session_state.get("alanlar", [])

    if alan_modu_aktif and alanlar_data:
        alanlar = [
            AlanTanimi(
                isim=a["isim"],
                gunluk_kontenjan=a.get("kontenjan", 1),
                max_kontenjan=a.get("max_kontenjan"),
                minimum_staffing=a.get("minimum_staffing", 1),
                kidem_kurallari=a.get("kidem_kurallari", {}),
                vardiya_tipleri=a.get("vardiya_tipleri", [])
            )
            for a in alanlar_data
        ]
    else:
        alanlar = []

    # Toplam hedef - zorunlu taban / teorik tavan bilgilendirmesi.
    # G1.3'ten beri hedefler soft: bu aralığın dışında olmak artık İMKÂNSIZ
    # DEĞİL, yalnızca sapma riskini işaret eder. st.stop() YOK — yapısal
    # eksiklik (örn. boş personel listesi) dışında çözüm engellenmez.
    toplam_hedef = sum(hedefler.values())
    enforce_minimum_staffing = st.session_state.get("enforce_minimum_staffing", True)
    hard_taban, teorik_tavan = _staffing_taban_tavan(
        alanlar, vardiyalar, gun_sayisi, enforce_minimum_staffing
    )

    if hard_taban > 0 and toplam_hedef < hard_taban:
        st.warning(
            f"Toplam hedef ({toplam_hedef}), zorunlu doluluk tabanının "
            f"({hard_taban}) altında. Plan yine üretilir; kişilere "
            f"hedeflerinden fazla nöbet düşecek ve sapmalar raporlanacak."
        )
    if teorik_tavan is not None and toplam_hedef > teorik_tavan:
        st.warning(
            f"Toplam hedef ({toplam_hedef}), teorik doluluk tavanının "
            f"({teorik_tavan}) üstünde. Kişiler hedeflerinin altında kalacak."
        )

    # Personel alan yetkinlikleri
    personel_alan_yetkinlikleri = st.session_state.get("personel_alan_yetkinlikleri", {})

    # Personel vardiya kısıtları
    personel_vardiya_kisitlari = st.session_state.get("personel_vardiya_kisitlari", {})

    # Solver config - kullanıcı ayarlarından al
    config = SolverConfig(
        # Hard constraints
        ardisik_yasak=st.session_state.get("ardisik_yasak", True),
        gunasiri_limit_aktif=st.session_state.get("gunasiri_limit_aktif", True),
        max_gunasiri_per_kisi=st.session_state.get("max_gunasiri", max_gunasiri_per_kisi),
        enforce_minimum_staffing=st.session_state.get("enforce_minimum_staffing", True),

        # Hafta sonu dengesi
        hafta_sonu_dengesi_aktif=st.session_state.get("hafta_sonu_dengesi", True),
        w_cuma=st.session_state.get("w_cuma", w_cuma),
        w_cumartesi=st.session_state.get("w_cumartesi", w_cumartesi),
        w_pazar=st.session_state.get("w_pazar", w_pazar),

        # Tatil dengesi
        tatil_dengesi_aktif=st.session_state.get("tatil_dengesi", True),

        # 2 gün boşluk tercihi
        iki_gun_bosluk_aktif=st.session_state.get("iki_gun_bosluk_aktif", True),
        w_iki_gun_bosluk=st.session_state.get("w_gap3", w_iki_gun_bosluk),

        # Saat bazlı denge
        saat_bazli_denge=st.session_state.get("saat_bazli_denge", True)
    )

    # Solver input
    # Önceki ay kuyruğunu oku (ay sınırı ardışık/dinlenme kısıtları için)
    onceki_kuyruk = onceki_ay_son_gun_atamalari(yil, ay, kac_gun=2)

    solver_input = SolverInput(
        yil=yil,
        ay=ay,
        personeller=personeller,
        hedefler=hedefler,
        vardiya_hedefleri=vardiya_hedefleri,
        izinler=izinler,
        tatiller=tatiller,
        ayri_tut=ayri_tut,
        birlikte_tut=birlikte_tut,
        esnek_ayri_tut=esnek_ayri_tut,
        tercih_edilen=tercih_edilen,
        alanlar=alanlar,
        personel_alan_yetkinlikleri=personel_alan_yetkinlikleri,
        alan_bazli_denklik=st.session_state.get("alan_bazli_denklik", True),
        personel_kidem_gruplari=st.session_state.get("personel_kidem_gruplari", {}),
        vardiyalar=vardiyalar,
        personel_vardiya_kisitlari=personel_vardiya_kisitlari,
        onceki_ay_kuyrugu=onceki_kuyruk,
        config=config
    )

    mod_bilgi = []
    if alanlar:
        mod_bilgi.append("Çoklu alan")
    if vardiyalar:
        mod_bilgi.append("Vardiya")
    if vardiya_hedefleri:
        mod_bilgi.append("Vardiya hedefleri")
    mod_str = f" ({', '.join(mod_bilgi)})" if mod_bilgi else ""
    st.info(f"Solver çalıştırılıyor...{mod_str}")

    solver = None
    try:
        solver = NobetSolver(solver_input)
        schedule = solver.coz()

        # Planı kaydet
        plan = AylikPlan(
            yil=yil,
            ay=ay,
            izinler={p: list(g) for p, g in izinler.items()},
            tercih_edilen_gunler={p: list(g) for p, g in tercih_edilen.items()},
            manuel_tatiller=list(manuel_holidays),
            hedef_override={p: h for p, h in hedefler.items() if h != default_target},
            sonuc={str(k): v for k, v in schedule.items()},
            sonuc_alanlı=bool(alanlar)
        )
        aylik_plani_kaydet(plan)

    except ValueError as e:
        # Bilinçli red: G1.2 pre-solve doğrulaması (örn. hedef > kişisel
        # maksimum) VEYA solver'ın INFEASIBLE raise'i (ikisi de ValueError).
        # Bu bir KURAL sorunu — teşhis anlamlı, çalıştırılır.
        meta = getattr(solver, "cozum_meta", None) if solver is not None else None
        baslik = "⚠️ Girdi kuralları çözümü engelliyor"
        if meta:
            baslik += f" (status: {meta['status']})"
        st.error(baslik)
        st.markdown(f"**{e}**")

        # Gelişmiş teşhis
        teshisler = gelismis_teshis(
            yil=yil,
            ay=ay,
            personeller=personeller,
            hedefler=hedefler,
            vardiya_hedefleri=vardiya_hedefleri,
            izinler=izinler,
            tatiller=tatiller,
            birlikte_tut=birlikte_tut,
            ayri_tut=ayri_tut,
            alanlar=alanlar if len(alanlar) > 0 else None,
            vardiyalar=vardiyalar if len(vardiyalar) > 0 else None,
            personel_alan_yetkinlikleri=personel_alan_yetkinlikleri,
            personel_vardiya_kisitlari=personel_vardiya_kisitlari,
            personel_kidem_gruplari=st.session_state.get("personel_kidem_gruplari", {}),
            ardisik_yasak=st.session_state.get("ardisik_yasak", True),
            enforce_minimum_staffing=st.session_state.get("enforce_minimum_staffing", True)
        )

        st.warning("🔍 **Tespit Edilen Sorunlar:**")

        errors = [t for t in teshisler if t.seviye == "error"]
        warnings = [t for t in teshisler if t.seviye == "warning"]

        if errors:
            st.markdown(f"**❌ {len(errors)} Kritik Sorun:**")
            for t in errors[:10]:
                with st.expander(f"🔴 {t.mesaj}", expanded=True):
                    st.json(t.detay)

        if warnings:
            st.markdown(f"**⚠️ {len(warnings)} Uyarı:**")
            for t in warnings[:5]:
                with st.expander(f"🟡 {t.mesaj}", expanded=False):
                    st.json(t.detay)

        st.stop()

    except Exception as e:
        # Kod hatası (TypeError, KeyError, AttributeError, ...) - bu bir
        # KURAL sorunu DEĞİL. Teşhis burada yanıltıcı olur, ÇALIŞTIRILMAZ.
        st.error("❌ Beklenmedik uygulama hatası")
        st.caption("Bu bir kural sorunu değil — geliştiriciye hata raporu iletebilirsiniz.")
        st.exception(e)
        st.stop()

    # Sonuç tablosu - mod'a göre farklı gösterim
    weekdays_tr = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]

    # Mod tespiti
    has_alanlar = bool(alanlar)
    has_vardiyalar = bool(vardiyalar)

    if has_alanlar and has_vardiyalar:
        # ALAN + VARDİYA MODU - {gun: {alan: {vardiya: [kişiler]}}}
        alan_isimleri = [a.isim for a in alanlar]
        vardiya_isimleri = [v.isim for v in vardiyalar]

        rows = []
        for gun in range(1, gun_sayisi + 1):
            dt = datetime(yil, ay, gun)
            wd = weekdays_tr[dt.weekday()]
            gun_data = schedule.get(gun, {})

            row = {
                "Gün": gun,
                "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                "Hafta Günü": wd,
                "Tatil": "Evet" if gun in tatiller else "",
            }

            # Her alan-vardiya kombinasyonu için sütun
            for alan_isim in alan_isimleri:
                alan_data = gun_data.get(alan_isim, {})
                for vardiya_isim in vardiya_isimleri:
                    kisiler = alan_data.get(vardiya_isim, [])
                    col_name = f"{alan_isim} / {vardiya_isim}"
                    row[col_name] = ", ".join(kisiler) if kisiler else "-"

            rows.append(row)

        df_schedule = pd.DataFrame(rows)

        st.success("🎉 Çözüm bulundu! (Çoklu Alan + Vardiya)")
        st.subheader("📋 Oluşturulan Nöbet Listesi")
        st.dataframe(df_schedule, use_container_width=True, hide_index=True)

        # İstatistikler
        st.divider()
        st.subheader("📊 Personel Dağılımı")

        stats = []
        for p in personeller:
            stat = {"Personel": p}
            toplam = 0
            toplam_saat = 0
            for gun_data in schedule.values():
                for alan_isim, alan_data in gun_data.items():
                    if isinstance(alan_data, dict):
                        for vardiya_isim, kisiler in alan_data.items():
                            if p in kisiler:
                                toplam += 1
                                # Saat hesapla
                                for v in vardiyalar:
                                    if v.isim == vardiya_isim:
                                        toplam_saat += v.saat
                                        break
            stat["Toplam Nöbet"] = toplam
            stat["Toplam Saat"] = toplam_saat
            stat["Hedef"] = hedefler.get(p, default_target)
            stats.append(stat)

        st.table(pd.DataFrame(stats))

    elif has_vardiyalar:
        # SADECE VARDİYA MODU - {gun: {vardiya: [kişiler]}}
        vardiya_isimleri = [v.isim for v in vardiyalar]

        rows = []
        for gun in range(1, gun_sayisi + 1):
            dt = datetime(yil, ay, gun)
            wd = weekdays_tr[dt.weekday()]
            gun_data = schedule.get(gun, {})

            row = {
                "Gün": gun,
                "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                "Hafta Günü": wd,
                "Tatil": "Evet" if gun in tatiller else "",
            }

            # Her vardiya için sütun
            for vardiya_isim in vardiya_isimleri:
                kisiler = gun_data.get(vardiya_isim, [])
                row[vardiya_isim] = ", ".join(kisiler) if kisiler else "-"

            rows.append(row)

        df_schedule = pd.DataFrame(rows)

        st.success("🎉 Çözüm bulundu! (Vardiya Modu)")
        st.subheader("📋 Oluşturulan Nöbet Listesi")
        st.dataframe(df_schedule, use_container_width=True, hide_index=True)

        # Vardiya bazlı dağılım istatistikleri
        st.divider()
        st.subheader("📊 Vardiya Bazlı Dağılım")

        stats = []
        for p in personeller:
            stat = {"Personel": p}
            toplam = 0
            toplam_saat = 0
            for vardiya in vardiyalar:
                count = sum(1 for g_data in schedule.values() if p in g_data.get(vardiya.isim, []))
                stat[vardiya.isim] = count
                toplam += count
                toplam_saat += count * vardiya.saat
            stat["TOPLAM"] = toplam
            stat["Saat"] = toplam_saat
            stat["Hedef"] = hedefler.get(p, default_target)
            stats.append(stat)

        st.table(pd.DataFrame(stats))

    elif has_alanlar:
        # ÇOKLU ALAN MODU - sonuç formatı: {gun: {alan: [kişiler]}}
        alan_isimleri = [a.isim for a in alanlar]

        rows = []
        for gun in range(1, gun_sayisi + 1):
            dt = datetime(yil, ay, gun)
            wd = weekdays_tr[dt.weekday()]
            gun_data = schedule.get(gun, {})

            row = {
                "Gün": gun,
                "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                "Hafta Günü": wd,
                "Tatil": "Evet" if gun in tatiller else "",
            }

            # Her alan için sütun
            for alan_isim in alan_isimleri:
                kisiler = gun_data.get(alan_isim, [])
                row[alan_isim] = ", ".join(kisiler) if kisiler else "-"

            rows.append(row)

        df_schedule = pd.DataFrame(rows)

        st.success("🎉 Çözüm bulundu! (Çoklu Alan Modu)")
        st.subheader("📋 Oluşturulan Nöbet Listesi")
        st.dataframe(df_schedule, use_container_width=True, hide_index=True)

        # Alan bazlı dağılım istatistikleri
        st.divider()
        st.subheader("📊 Alan Bazlı Dağılım")

        alan_stats = []
        for p in personeller:
            stat = {"Personel": p}
            toplam = 0
            for alan_isim in alan_isimleri:
                count = sum(1 for g in schedule.values() if p in g.get(alan_isim, []))
                stat[alan_isim] = count
                toplam += count
            stat["TOPLAM"] = toplam
            stat["Hedef"] = hedefler.get(p, default_target)
            alan_stats.append(stat)

        st.table(pd.DataFrame(alan_stats))

    else:
        # TEK ALAN MODU - eski format: {gun: [kişiler]}
        max_kisi = max((len(v) for v in schedule.values() if isinstance(v, list)), default=1)

        rows = []
        for gun in range(1, gun_sayisi + 1):
            dt = datetime(yil, ay, gun)
            wd = weekdays_tr[dt.weekday()]
            isimler = schedule.get(gun, [])
            if not isinstance(isimler, list):
                isimler = []
            row = {
                "Gün": gun,
                "Tarih": f"{gun:02d}/{ay:02d}/{yil}",
                "Hafta Günü": wd,
                "Kişi Sayısı": len(isimler),
                "Tatil": "Evet" if gun in tatiller else "",
            }
            for i in range(max_kisi):
                row[f"Nöbetçi {i+1}"] = isimler[i] if i < len(isimler) else ""
            rows.append(row)

        df_schedule = pd.DataFrame(rows)

        st.success("🎉 Çözüm bulundu!")
        st.subheader("📋 Oluşturulan Nöbet Listesi")
        st.dataframe(df_schedule, use_container_width=True, hide_index=True)

        # Personel dağılımı
        st.divider()
        st.subheader("📊 Personel Nöbet Dağılımı")
        stats = []
        for p in personeller:
            count = sum(1 for v in schedule.values() if isinstance(v, list) and p in v)
            hedef = hedefler.get(p, default_target)
            stats.append({
                "Personel": p,
                "Hedef": hedef,
                "Gerçekleşen": count,
                "Fark": count - hedef
            })

        st.table(pd.DataFrame(stats))

    # CSV indirme (her iki mod için)
    csv_data = df_schedule.to_csv(index=False).encode('utf-8-sig')
    st.download_button(
        "📥 CSV İndir",
        data=csv_data,
        file_name=f"Nobet_{yil}_{ay:02d}.csv",
        mime="text/csv"
    )

    # Excel indirme
    xlsx_buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nöbet {ay:02d}-{yil}"

    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    fieldnames = list(rows[0].keys())
    for c, h in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    fill_weekend = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    fill_holiday = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    for r_i, row in enumerate(rows, start=2):
        dt = datetime(yil, ay, row["Gün"])
        is_weekend = weekdays_tr[dt.weekday()] in ["Cuma", "Cumartesi", "Pazar"]
        is_holiday = row["Gün"] in tatiller

        for c_i, h in enumerate(fieldnames, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=row.get(h, ""))
            if c_i <= 5:
                cell.alignment = center

            if is_holiday:
                cell.fill = fill_holiday
            elif is_weekend:
                cell.fill = fill_weekend

    for col in ws.columns:
        maxlen = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 30)

    wb.save(xlsx_buf)
    xlsx_buf.seek(0)

    st.download_button(
        "⬇️ Excel İndir (XLSX)",
        data=xlsx_buf.getvalue(),
        file_name=f"nobet_{ay:02d}_{yil}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
